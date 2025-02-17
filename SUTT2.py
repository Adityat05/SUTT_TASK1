import openpyxl
import json
path="Mess Menu Sample.xlsx" #path of file
wb_obj=openpyxl.load_workbook(path)
sheet_obj=wb_obj.active 
row_count = sheet_obj.max_row #take number of rows      
column_count=sheet_obj.max_column #take number of columns
#print(row_count)
# Rows=row_count
# Columns=column_count
Dict_final={}
Days=["MONDAY","TUESDAY","WEDNESDAY","THURSDAY","FRIDAY","SATURDAY","SUNDAY"]
Meals=["BREAKFAST","LUNCH","DINNER"]
for i in range(1,column_count): #to iterate through all the columns
    Dict_col={} #the dictionary for each column
    DATe=""
    B=[]
    L=[]
    D=[]
    flag=0
    temp_col=sheet_obj.cell(row=1, column=i) #to check if we are iterating on an empty column
    if (temp_col=="" or temp_col is None):
        continue
    for j in range(2,row_count): #iterating through the current column
        temp=sheet_obj.cell(row=j, column=i).value
        if (j==2): #as the date is present on the second row only in all the columns
            DATe=temp
            continue
        if (temp in Days or temp in Meals): #to check on which Meal we are iterating 
            if (temp in Meals):
                if (temp=="BREAKFAST"):
                    flag=0
                elif (temp=="LUNCH"):
                    flag=1
                else:
                    flag=2
            if (temp in Days):
                continue
        if (temp=="" or temp is None): #to check empty cells
            continue
        if (temp[0]=="*"): #to remove '***' type elements
            continue
        if (flag==0):
            B.append(temp)
        if (flag==1):
            L.append(temp)      
        if (flag==2):
            D.append(temp)
    Dict_col["BREAKFAST"]=B
    Dict_col["LUNCH"]=L
    Dict_col["DINNER"]=D
    DATe=DATe.strftime("%Y-%m-%d") #formatting the Date in actual usable format
    Dict_final[DATe]=Dict_col
#print(Dict_final)
with open('Data.json','w') as f:
    json.dump(Dict_final, f, indent=4, sort_keys=False) #pretty json we keep indent 4, it indents itself on each new List or dictionary element
    print("JSON FILE CREATED SUCCESSFULLY")
